#Project: Toner Status Update
#File: Excel Write
#Name: Trace Watkins
#Date: 4/29/2019

from ast import literal_eval
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles.alignment import Alignment
import regex as re


#declaring global variables
ip_not_found = True
status_not_found = True

#insert in excel function
def excel_write(sheet):
    #find column with IP addresses
    for cols in sheet.iter_cols(min_row=1, max_col=10, max_row=1):
        if cols[0].value == "IP":
            ip_not_found = False
            col_num = cols[0].col_idx
        if cols[0].value == 'Status':
            status_not_found = False
            stat_col = cols[0].col_idx

    #Notify if IP column not found    
    if ip_not_found == True:
        print('Could not locate IP title in first row')
    if status_not_found == True:
        print('Could not locate Status title in first row')


    #iterate through IPs and insert value
    for rows in sheet.iter_rows(min_col=col_num, min_row=2, max_col=col_num, max_row=50):
        edit_ip = re.sub(r'[\n\t\s\r]+','',str(rows[0].value))
        if edit_ip in new_dict:

            c = sheet.cell(row=rows[0].row, column=stat_col)
            c.value = new_dict[edit_ip]
            c.alignment = Alignment(horizontal='center')



#read string from ip_dict.txt as dict
f = open("C:\\ip_dict.txt","r")
new_dict = literal_eval(f.read())
f.close()

#load printer workbook
wb = load_workbook(filename ="T:\Printer List (Updated).xlsx")
branch = wb['Branch Printers']
corp = wb['CorpOps']

excel_write(branch)
excel_write(corp)

wb.save(filename ="T:\Printer List (Updated).xlsx")
